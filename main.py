import re
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# Get a logger instance
logger = logging.getLogger(__name__)

# Create the output directory
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

file_path = "data/ITC MATCH.xlsx"

# Load data
sheets = ["GSTN", "BOOKS"]
data = {
    sheet: pd.read_excel(file_path, sheet_name=sheet, skiprows=1) for sheet in sheets
}
logger.info("File loaded, processing ...")


def save_to_excel(file_path, result):
    try:
        # Load existing workbook
        wb = load_workbook(file_path)

        for sheet_name, df in result.items():
            logger.info(f"Saving {sheet_name} sheet...")
            if df is not None and not df.empty:
                df = clean_columns(df, sheet_name)

                # If sheet doesn't exist, create it
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                ws = wb[sheet_name]

                # Remove existing data starting from row 3 (if any)
                max_col = ws.max_column
                max_row = ws.max_row
                if max_row >= 3:
                    ws.delete_rows(3, max_row - 2)

                # Write dataframe rows to worksheet starting from row 3
                for r_idx, row in enumerate(
                    dataframe_to_rows(df, index=False, header=False), start=3
                ):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

        # Save the workbook
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Error in saving sheets : {e}")
        return False


def clean_columns(df: pd.DataFrame, sheet_name: str):
    cols = df.columns
    clean_cols = []
    for col in cols:
        if col in [
            "InvoiceNumber_clean",
            "InvoiceNumber_original_gstn",
            "InvoiceNumber_original_books",
            "_merge",
        ]:
            df.drop(columns=col, inplace=True)
            continue
        elif col.endswith("_books"):
            if sheet_name == "MATCHED":
                df.drop(columns=col, inplace=True)
                continue
            new_col = col.replace("_books", "")
        elif col.endswith("_gstn"):
            if sheet_name == "NEXT_FY_ITC":
                df.drop(columns=col, inplace=True)
                continue
            new_col = col.replace("_gstn", "")
        else:
            if col == "gstr1_filing_date":
                if sheet_name != "MATCHED":
                    df.drop(columns=col, inplace=True)
                    continue
            new_col = col
        clean_cols.append(new_col)
    df.columns = clean_cols

    # Format Date
    if "Invoice Date" in df.columns:
        df["Invoice Date"] = pd.to_datetime(
            df["Invoice Date"], errors="coerce"
        ).dt.strftime("%d/%m/%Y")
    if "gstr1_filing_date" in df.columns:
        df["gstr1_filing_date"] = pd.to_datetime(
            df["gstr1_filing_date"], errors="coerce"
        ).dt.strftime("%d/%m/%Y")

    return df


def clean_invoice_number(invoice_str):
    """Remove only alphabets while preserving all other characters"""
    if pd.isna(invoice_str):
        return None
    # Remove only alphabets (both lowercase and uppercase)
    return re.sub(r"[a-zA-Z]", "", str(invoice_str))


# Clean data while preserving original format (minus alphabets)
gstn = (
    data["GSTN"]
    .dropna(subset=["GSTN", "Invoice Number"])
    .assign(
        InvoiceNumber_clean=lambda x: x["Invoice Number"].apply(clean_invoice_number),
        InvoiceNumber_original=lambda x: x["Invoice Number"],
    )
    .dropna(subset=["InvoiceNumber_clean"])
)

books = (
    data["BOOKS"]
    .dropna(subset=["GSTN", "Invoice Number"])
    .assign(
        InvoiceNumber_clean=lambda x: x["Invoice Number"].apply(clean_invoice_number),
        InvoiceNumber_original=lambda x: x["Invoice Number"],
    )
    .dropna(subset=["InvoiceNumber_clean"])
)

# Merge data
merged = pd.merge(
    gstn,
    books,
    on=["GSTN", "InvoiceNumber_clean"],
    how="outer",
    indicator=True,
    suffixes=("_gstn", "_books"),  # Add suffix only to books columns
)

# For matched records
matched_df = merged[merged["_merge"] == "both"].copy()


def add_mismatch_flag(df):
    """Add mismatch flag for tax values"""
    tax_columns = [
        "Taxable",
        "CGST",
        "SGST",
        "IGST",
        "CESS",
    ]  # Adjusted column names

    # Ensure tax columns exist in both GSTN and BOOKS data
    for col in tax_columns:
        if f"{col}_gstn" not in df.columns or f"{col}_books" not in df.columns:
            print(f"Warning: Missing tax column {col} in merged data")
            continue

        df[f"{col}_Match"] = df[f"{col}_gstn"].fillna(0).astype(float) == df[
            f"{col}_books"
        ].fillna(0).astype(float)

    # Create overall mismatch flag
    match_cols = [f"{col}_Match" for col in tax_columns if f"{col}_Match" in df.columns]
    if match_cols:
        df["MIS_MATCHED"] = ~df[match_cols].all(axis=1)
        df = df.drop(columns=match_cols)

    return df


# Add mismatch flags
matched_df = add_mismatch_flag(matched_df)


def categorize_gstn(df, cutoff_date="2024-04-01"):
    """Categorize GSTN records by fiscal year"""
    prev_fy = []
    not_in_books = []

    for _, row in df.iterrows():
        inv_date = row.get("Invoice Date_gstn")
        if pd.notnull(inv_date):
            inv_date = pd.to_datetime(inv_date)
            if inv_date < pd.Timestamp(cutoff_date):
                prev_fy.append(row)
            else:
                not_in_books.append(row)
        else:
            print(
                f"Missing Invoice Date: {row['GSTN']}, {row['InvoiceNumber_original_gstn']}"
            )
            not_in_books.append(row)

    return pd.DataFrame(prev_fy), pd.DataFrame(not_in_books)


# Prepare final results
result = {
    "MATCHED": matched_df.drop(columns="_merge"),
    "PREV_FY_ITC": None,
    "NOTINBOOKS": None,
    "NEXT_FY_ITC": None,
}


# Update NEXT_FY_ITC sheet
next_fy_df = merged[merged["_merge"] == "right_only"]
next_fy_df = next_fy_df.filter(
    regex="GSTN|InvoiceNumber_original_books|^((?!_clean|_gstn).)*$"
)
# Move 'gstr1_filing_date' to the end if it exists
if "gstr1_filing_date" in next_fy_df.columns:
    cols = [col for col in next_fy_df.columns if col != "gstr1_filing_date"]
    cols.append("gstr1_filing_date")
    next_fy_df = next_fy_df[cols]

result["NEXT_FY_ITC"] = next_fy_df


# Categorize unmatched GSTN records
gstn_unmatched = merged[merged["_merge"] == "left_only"].filter(
    regex="GSTN|InvoiceNumber_original_gstn|^((?!_books).)*$"
)

result["PREV_FY_ITC"], result["NOTINBOOKS"] = categorize_gstn(gstn_unmatched)

# Export to Excel
if not save_to_excel(file_path, result):
    logger.error(f"Error in writing sheets to the file : {file_path}")

logger.info(f"Processing successfully completed.")
logger.info(f"File saved to the path : {file_path}.")

# # Export to Excel
# with pd.ExcelWriter(
#     file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
# ) as writer:
#     for sheet_name, df in result.items():
#         if df is not None and not df.empty:
#             # Clean column names for output
#             df = clean_columns(df, sheet_name)
#             df.to_excel(writer, sheet_name=sheet_name, index=False)


# # Export to Excel
# out_file_path = os.path.join(OUTPUT_DIR, os.path.basename(file_path))
# with pd.ExcelWriter(out_file_path, engine="openpyxl") as writer:
#     for sheet_name, df in result.items():
#         if df is not None and not df.empty:
#             # Clean column names for output
#             df = clean_columns(df, sheet_name)
#             df.to_excel(writer, sheet_name=sheet_name, index=False)
